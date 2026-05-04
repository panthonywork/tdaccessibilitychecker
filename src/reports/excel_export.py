from io import BytesIO


def build_excel(results: list[dict], filename: str) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "Accessibility Issues"

    headers = ["Checkpoint ID", "Group", "Status", "Issue", "How to Fix"]
    header_fill = PatternFill("solid", fgColor="004C97")
    header_font = Font(bold=True, color="FFFFFF")

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True)

    status_fills = {
        "pass": PatternFill("solid", fgColor="C6EFCE"),
        "fail": PatternFill("solid", fgColor="FFC7CE"),
        "na":   PatternFill("solid", fgColor="E0E0E0"),
    }
    status_labels = {"pass": "PASS", "fail": "FAIL", "na": "N/A"}

    for row_num, result in enumerate(results, 2):
        status = result.get("status", "na")
        ws.cell(row=row_num, column=1, value=result.get("id", ""))
        ws.cell(row=row_num, column=2, value=result.get("group_name", ""))
        ws.cell(row=row_num, column=3, value=status_labels.get(status, status.upper()))
        ws.cell(row=row_num, column=4, value=result.get("plain_english", ""))
        ws.cell(row=row_num, column=5, value=result.get("fix_hint", "") if status == "fail" else "")

        fill = status_fills.get(status, status_fills["na"])
        for col in range(1, 6):
            ws.cell(row=row_num, column=col).fill = fill
            ws.cell(row=row_num, column=col).alignment = Alignment(wrap_text=True)

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 55
    ws.column_dimensions["E"].width = 60

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
